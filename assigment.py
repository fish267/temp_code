# 先执行 pip install openpyxl
import openpyxl
import datetime

# 读取考试成绩表
score_wb = openpyxl.load_workbook('exam_score.xlsx')
score_sheet = score_wb.active

# 读取志愿表
aspiration_wb = openpyxl.load_workbook('aspiration.xlsx')
aspiration_sheet = aspiration_wb.active

# 创建录取结果列表和占用名额计数器
admission_result = []
admission_count = 0
admission_limit = {}  # 志愿的名额限制

# 获取志愿表中的名额限制
for row in aspiration_sheet.iter_rows(min_row=2, values_only=True):
    aspiration = row[0]
    limit = row[1]
    admission_limit[aspiration] = limit

TOTAL = sum(admission_limit.values())
# 按照第一志愿的分数由高到低排序考生
sorted_score_rows = sorted(score_sheet.iter_rows(min_row=2, values_only=True), key=lambda x: x[6], reverse=True)

# 遍历每个考生
for score_row in sorted_score_rows:
    student_number = score_row[0]
    student_name = score_row[1]
    total_score = score_row[6]  # 获取总分
    agree_transfer = score_row[5]
    admission_complete = False

    # 如果录取人数超过名额限制，则后续的考生都标记为未录取
    start = 0
    for result in admission_result:
        if result[-1] == '成功录取':
            start += 1
    if start >= TOTAL:
        admission_result.append((student_number, student_name, total_score, "", "未录取",))

    else:
        # 遍历考生的志愿，按照指定优先级进行考虑
        aspirations = [score_row[2], score_row[3], score_row[4]]
        for aspiration in aspirations:
            # 忽略空志愿
            if not aspiration:
                continue

            if admission_complete:
                break

            admission_limit_aspiration = admission_limit.get(aspiration, 0)
            if admission_limit_aspiration > 0 and admission_count < admission_limit_aspiration:
                admission_limit_aspiration -= 1
                admission_result.append((student_number, student_name, total_score, aspiration, "成功录取",))
                admission_complete = True
                admission_limit[aspiration] = admission_limit_aspiration
                break

        # 如果考生没有录取到志愿中，则标记为未录取
        if not admission_complete:
            if agree_transfer == '同意':
                admission_result.append((student_number, student_name, total_score, "待分配", "成功录取",))
            else:
                admission_result.append((student_number, student_name, total_score, "", "未录取",))

# 打印录取结果
for result in admission_result:
    student_number, student_name, total_score, aspiration, admission_desc, = result
    print('考生号：', student_number)
    print('姓名：', student_name)
    print('总分：', total_score)

    print('录取志愿：', aspiration)
    print('录取状态：', admission_desc)

    print('---------------------')

# 创建新的工作簿和工作表
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active

# 复制标题行到新的工作表
output_sheet.append(['考生号', '姓名', '总分', '录取志愿', '录取状态'])

# 复制更新后的成绩到新的工作表
for result in admission_result:
    output_sheet.append(result)

# 添加时间戳
timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
output_file = f'updated_scores_{timestamp}.xlsx'

# 保存更新后的成绩到新的文件
output_wb.save(output_file)
